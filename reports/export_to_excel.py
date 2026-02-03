"""
SHUNCOM RULR Feature Requirements Checklist - Excel Exporter
Enterprise-grade export with comprehensive tracking columns
"""

import re
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Read the markdown file
with open(r'd:\OneDrive - linhdangdev\Documents\Obsidian Vault\ShucomAIOT\SHUNCOM RULR Knowledge Base\Feature Requirements Checklist.md', 'r', encoding='utf-8') as f:
    content = f.read()

# Parse the markdown content
features = []
current_phase = ""
current_module = ""
current_category = ""
current_feature = ""
feature_id_counter = 0

lines = content.split('\n')

for line in lines:
    # Detect Phase
    if line.startswith('## 🔐 Phase 1:') or line.startswith('## 🔧 Phase 1:') or line.startswith('## ⚙️ Phase 1:') or line.startswith('## 📱 Phase 1:'):
        current_phase = "Phase 1: Core Infrastructure"
        current_module = line.replace('## ', '').replace('🔐 ', '').replace('🔧 ', '').replace('⚙️ ', '').replace('📱 ', '').strip()
    elif line.startswith('## 🗺️ Phase 2:') or line.startswith('## 🔄 Phase 2:') or line.startswith('## 📊 Phase 2:'):
        current_phase = "Phase 2: Advanced Features"
        current_module = line.replace('## ', '').replace('🗺️ ', '').replace('🔄 ', '').replace('📊 ', '').strip()
    elif line.startswith('## 📈 Phase 3:'):
        current_phase = "Phase 3: Analytics & Optimization"
        current_module = line.replace('## ', '').replace('📈 ', '').strip()
    elif line.startswith('## 🚀'):
        current_phase = "Cross-Phase"
        current_module = "Performance & Optimization"
    elif line.startswith('## 🧪'):
        current_phase = "Cross-Phase"
        current_module = "Testing & Quality Assurance"
    elif line.startswith('## 📱 Mobile'):
        current_phase = "Cross-Phase"
        current_module = "Mobile & Responsive Design"
    elif line.startswith('## 🔗'):
        current_phase = "Cross-Phase"
        current_module = "Integration & API"
    elif line.startswith('## ⚙️ DevOps'):
        current_phase = "Cross-Phase"
        current_module = "DevOps & Deployment"
    elif line.startswith('## 📋 Project'):
        current_phase = "Cross-Phase"
        current_module = "Project Management"

    # Detect Category (### level)
    if line.startswith('### '):
        current_category = line.replace('### ', '').strip()

    # Detect Feature Group (bold items with checkbox)
    feature_match = re.match(r'^- \[ \] \*\*(.+)\*\*$', line)
    if feature_match:
        current_feature = feature_match.group(1)

    # Detect Sub-features (indented checkboxes)
    sub_feature_match = re.match(r'^  - \[ \] (.+)$', line)
    if sub_feature_match and current_phase:
        feature_id_counter += 1
        feature_name = sub_feature_match.group(1)

        # Determine priority based on phase
        if "Phase 1" in current_phase:
            priority = "Critical"
        elif "Phase 2" in current_phase:
            priority = "High"
        elif "Phase 3" in current_phase:
            priority = "Medium"
        else:
            priority = "Standard"

        features.append({
            'id': f'FR-{feature_id_counter:04d}',
            'phase': current_phase,
            'module': current_module,
            'category': current_category,
            'feature_group': current_feature,
            'feature_name': feature_name,
            'priority': priority,
            'status': 'Not Started',
            'assignee': '',
            'estimated_hours': '',
            'actual_hours': '',
            'start_date': '',
            'due_date': '',
            'completion_date': '',
            'notes': '',
            'blockers': '',
            'dependencies': '',
            'acceptance_criteria': '',
            'test_status': 'Not Tested'
        })

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Feature Requirements"

# Define headers
headers = [
    'Feature ID', 'Phase', 'Module', 'Category', 'Feature Group', 'Feature Name',
    'Priority', 'Status', 'Assignee', 'Est. Hours', 'Actual Hours',
    'Start Date', 'Due Date', 'Completion Date', 'Notes', 'Blockers',
    'Dependencies', 'Acceptance Criteria', 'Test Status'
]

# Style definitions
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Priority colors
priority_fills = {
    'Critical': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),
    'High': PatternFill(start_color='FFB347', end_color='FFB347', fill_type='solid'),
    'Medium': PatternFill(start_color='77DD77', end_color='77DD77', fill_type='solid'),
    'Standard': PatternFill(start_color='89CFF0', end_color='89CFF0', fill_type='solid')
}

# Status colors
status_fills = {
    'Not Started': PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid'),
    'In Progress': PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid'),
    'Completed': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),
    'Blocked': PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'),
    'On Hold': PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
}

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Write data
for row, feature in enumerate(features, 2):
    ws.cell(row=row, column=1, value=feature['id']).border = thin_border
    ws.cell(row=row, column=2, value=feature['phase']).border = thin_border
    ws.cell(row=row, column=3, value=feature['module']).border = thin_border
    ws.cell(row=row, column=4, value=feature['category']).border = thin_border
    ws.cell(row=row, column=5, value=feature['feature_group']).border = thin_border
    ws.cell(row=row, column=6, value=feature['feature_name']).border = thin_border

    # Priority with color
    priority_cell = ws.cell(row=row, column=7, value=feature['priority'])
    priority_cell.border = thin_border
    priority_cell.fill = priority_fills.get(feature['priority'], PatternFill())

    # Status with color
    status_cell = ws.cell(row=row, column=8, value=feature['status'])
    status_cell.border = thin_border
    status_cell.fill = status_fills.get(feature['status'], PatternFill())

    # Remaining columns
    for col in range(9, 20):
        cell = ws.cell(row=row, column=col, value='')
        cell.border = thin_border

# Auto-adjust column widths
column_widths = {
    1: 12,   # Feature ID
    2: 25,   # Phase
    3: 30,   # Module
    4: 25,   # Category
    5: 25,   # Feature Group
    6: 45,   # Feature Name
    7: 12,   # Priority
    8: 15,   # Status
    9: 15,   # Assignee
    10: 12,  # Est. Hours
    11: 12,  # Actual Hours
    12: 12,  # Start Date
    13: 12,  # Due Date
    14: 15,  # Completion Date
    15: 30,  # Notes
    16: 20,  # Blockers
    17: 20,  # Dependencies
    18: 40,  # Acceptance Criteria
    19: 15   # Test Status
}

for col, width in column_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# Freeze header row
ws.freeze_panes = 'A2'

# Add Summary Sheet
summary = wb.create_sheet(title="Summary Dashboard")

# Summary headers
summary_headers = ['Metric', 'Value']
for col, header in enumerate(summary_headers, 1):
    cell = summary.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

# Calculate metrics
phase_counts = {}
priority_counts = {}
for f in features:
    phase_counts[f['phase']] = phase_counts.get(f['phase'], 0) + 1
    priority_counts[f['priority']] = priority_counts.get(f['priority'], 0) + 1

summary_data = [
    ('Total Features', len(features)),
    ('', ''),
    ('--- BY PHASE ---', ''),
]

for phase, count in phase_counts.items():
    summary_data.append((phase, count))

summary_data.extend([
    ('', ''),
    ('--- BY PRIORITY ---', ''),
])

for priority, count in priority_counts.items():
    summary_data.append((priority, count))

summary_data.extend([
    ('', ''),
    ('--- PROJECT INFO ---', ''),
    ('Generated Date', datetime.now().strftime('%Y-%m-%d %H:%M')),
    ('Project', 'SHUNCOM RULR IoT Platform'),
    ('Company', 'Shanghai Shuncom AIOT Co., Ltd'),
])

for row, (metric, value) in enumerate(summary_data, 2):
    summary.cell(row=row, column=1, value=metric).border = thin_border
    summary.cell(row=row, column=2, value=value).border = thin_border

summary.column_dimensions['A'].width = 30
summary.column_dimensions['B'].width = 40

# Add Data Validation Sheet (Status options)
validation = wb.create_sheet(title="Validation Lists")
status_options = ['Not Started', 'In Progress', 'Completed', 'Blocked', 'On Hold', 'Under Review']
priority_options = ['Critical', 'High', 'Medium', 'Low', 'Standard']
test_options = ['Not Tested', 'Testing', 'Passed', 'Failed', 'Skipped']

validation['A1'] = 'Status Options'
validation['B1'] = 'Priority Options'
validation['C1'] = 'Test Status Options'

for i, status in enumerate(status_options, 2):
    validation[f'A{i}'] = status

for i, priority in enumerate(priority_options, 2):
    validation[f'B{i}'] = priority

for i, test in enumerate(test_options, 2):
    validation[f'C{i}'] = test

# Save workbook
output_path = r'd:\OneDrive - linhdangdev\Documents\Obsidian Vault\ShucomAIOT\reports\SHUNCOM_RULR_Feature_Checklist.xlsx'
wb.save(output_path)

print(f"\n✅ Excel file exported successfully!")
print(f"📁 Location: {output_path}")
print(f"📊 Total features exported: {len(features)}")
print(f"\n--- Phase Distribution ---")
for phase, count in phase_counts.items():
    print(f"  • {phase}: {count}")
print(f"\n--- Priority Distribution ---")
for priority, count in priority_counts.items():
    print(f"  • {priority}: {count}")
